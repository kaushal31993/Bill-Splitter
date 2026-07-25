"""Excel export.

Two design choices worth stating, because they are what make the workbook
useful rather than just a picture of the numbers:

* Money is written as **real numeric cells** with a currency number format, not
  pre-formatted strings, so the recipient can sum, sort, and pivot them.
* Column totals are written as **live SUM formulas**, so the workbook
  demonstrates its own arithmetic instead of asking the reader to trust it.
"""

from __future__ import annotations

import io
import re
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import models, schemas
from .money import to_dollars

CURRENCY_FORMAT = '"$"#,##0.00;[Red]-"$"#,##0.00'
DATE_FORMAT = "MM/DD/YYYY"

HEADER_FILL = PatternFill("solid", fgColor="1F3B4D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
TOP_BORDER = Border(top=Side(style="thin", color="808080"))


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: 31 chars max, and []:*?/\\ are illegal."""
    cleaned = re.sub(r"[\[\]:*?/\\]", "-", name).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 2
    while candidate.lower() in used:
        tail = f" ({suffix})"
        candidate = cleaned[: 31 - len(tail)] + tail
        suffix += 1
    used.add(candidate.lower())
    return candidate


def _money_cell(ws, row: int, col: int, cents: int):
    cell = ws.cell(row=row, column=col, value=to_dollars(cents))
    cell.number_format = CURRENCY_FORMAT
    return cell


def _header_row(ws, row: int, values: list[str]) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_workbook(
    event: models.Event, totals: schemas.EventTotalsOut
) -> bytes:
    participants = list(event.participants)
    names = {p.id: p.display_name for p in participants}

    wb = Workbook()
    used_names: set[str] = set()

    _build_summary(wb.active, event, totals, participants, names, used_names)

    breakdowns = {b.bill_id: b for b in totals.bills}
    for bill in event.bills:
        bd = breakdowns.get(bill.id)
        if bd is None:
            continue
        title = _sanitize_sheet_name(bill.label, used_names)
        ws = wb.create_sheet(title=title)
        _build_bill_sheet(ws, bill, bd, participants, names)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary(ws, event, totals, participants, names, used_names) -> None:
    ws.title = _sanitize_sheet_name("Summary", used_names)

    ws["A1"] = event.name
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Bill Splitter summary"
    ws["A2"].font = Font(italic=True, color="666666")

    ws["A4"] = "Event total"
    ws["A4"].font = TOTAL_FONT
    _money_cell(ws, 4, 2, totals.grand_total_cents).font = TOTAL_FONT

    if not totals.is_complete:
        ws["A5"] = "Incomplete: some items are unassigned or a payer is not set."
        ws["A5"].font = Font(italic=True, color="B03A2E")

    header_row = 7
    _header_row(ws, header_row, ["Person", "Owes", "Paid", "Net"])

    first = header_row + 1
    for offset, p in enumerate(participants):
        row = first + offset
        ws.cell(row=row, column=1, value=names[p.id])
        _money_cell(ws, row, 2, totals.totals_cents.get(p.id, 0))
        _money_cell(ws, row, 3, totals.paid_cents.get(p.id, 0))
        net = totals.net_cents.get(p.id, 0)
        cell = _money_cell(ws, row, 4, net)
        # D = B - C, written live so the relationship is visible in the sheet.
        cell.value = f"=B{row}-C{row}"

    last = first + len(participants) - 1
    total_row = last + 1 if participants else first
    ws.cell(row=total_row, column=1, value="Total").font = TOTAL_FONT
    for col in (2, 3, 4):
        if participants:
            cell = ws.cell(
                row=total_row,
                column=col,
                value=f"=SUM({get_column_letter(col)}{first}:{get_column_letter(col)}{last})",
            )
        else:
            cell = ws.cell(row=total_row, column=col, value=Decimal("0.00"))
        cell.number_format = CURRENCY_FORMAT
        cell.font = TOTAL_FONT
        cell.border = TOP_BORDER

    row = total_row + 2
    ws.cell(row=row, column=1, value="Who owes whom").font = TOTAL_FONT
    row += 1
    if totals.debts:
        _header_row(ws, row, ["From", "To", "Amount"])
        row += 1
        for debt in totals.debts:
            ws.cell(row=row, column=1, value=names.get(debt.from_participant_id, "?"))
            ws.cell(row=row, column=2, value=names.get(debt.to_participant_id, "?"))
            _money_cell(ws, row, 3, debt.amount_cents)
            row += 1
    else:
        ws.cell(row=row, column=1, value="Everything is settled.").font = Font(italic=True)

    _autosize(ws, {1: 28, 2: 14, 3: 14, 4: 14})
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _build_bill_sheet(ws, bill, bd, participants, names) -> None:
    ws["A1"] = bill.label + (f" — {bill.merchant}" if bill.merchant else "")
    ws["A1"].font = TITLE_FONT

    meta_row = 2
    if bill.bill_date:
        ws.cell(row=meta_row, column=1, value="Date")
        cell = ws.cell(row=meta_row, column=2, value=bill.bill_date)
        cell.number_format = DATE_FORMAT
        meta_row += 1
    ws.cell(row=meta_row, column=1, value="Paid by")
    ws.cell(row=meta_row, column=2, value=names.get(bd.payer_id, "— not set —"))

    header_row = meta_row + 2
    headers = ["Item", "Qty", "Line total"] + [names[p.id] for p in participants]
    _header_row(ws, header_row, headers)

    first_item_row = header_row + 1
    row = first_item_row
    for item in bill.items:
        ws.cell(row=row, column=1, value=item.name or "(unnamed)")
        ws.cell(row=row, column=2, value=item.quantity)
        _money_cell(ws, row, 3, item.total_cents)
        shares = bd.item_shares.get(item.id, {})
        if not shares:
            ws.cell(row=row, column=1).font = Font(color="B03A2E")
            ws.cell(row=row, column=1).comment = None
        for col_offset, p in enumerate(participants):
            _money_cell(ws, row, 4 + col_offset, shares.get(p.id, 0))
        row += 1

    last_item_row = row - 1
    has_items = last_item_row >= first_item_row

    # Subtotal line, then the charge allocations, then the per-person total.
    share_by_pid = {s.participant_id: s for s in bd.shares}

    def charge_row(r: int, label: str, amount_cents: int, attr: str, negate: bool = False) -> None:
        ws.cell(row=r, column=1, value=label).font = TOTAL_FONT
        _money_cell(ws, r, 3, -amount_cents if negate else amount_cents)
        for col_offset, p in enumerate(participants):
            share = share_by_pid.get(p.id)
            value = getattr(share, attr) if share else 0
            _money_cell(ws, r, 4 + col_offset, -value if negate else value)

    subtotal_row = row
    ws.cell(row=subtotal_row, column=1, value="Items subtotal").font = TOTAL_FONT
    for col in [3] + [4 + i for i in range(len(participants))]:
        letter = get_column_letter(col)
        cell = ws.cell(
            row=subtotal_row,
            column=col,
            value=(
                f"=SUM({letter}{first_item_row}:{letter}{last_item_row})"
                if has_items
                else Decimal("0.00")
            ),
        )
        cell.number_format = CURRENCY_FORMAT
        cell.font = TOTAL_FONT
        cell.border = TOP_BORDER

    row = subtotal_row + 1
    charge_row(row, "Tax", bill.tax_cents, "tax_cents")
    row += 1
    charge_row(row, "Tip", bill.tip_cents, "tip_cents")
    row += 1
    charge_row(row, "Fees", bill.fee_cents, "fee_cents")
    row += 1
    charge_row(row, "Discount", bill.discount_cents, "discount_cents", negate=True)
    charges_first, charges_last = subtotal_row, row

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    for col in [3] + [4 + i for i in range(len(participants))]:
        letter = get_column_letter(col)
        cell = ws.cell(
            row=row, column=col, value=f"=SUM({letter}{charges_first}:{letter}{charges_last})"
        )
        cell.number_format = CURRENCY_FORMAT
        cell.font = TOTAL_FONT
        cell.border = TOP_BORDER

    widths = {1: 34, 2: 6, 3: 13}
    for i in range(len(participants)):
        widths[4 + i] = 14
    _autosize(ws, widths)
    ws.freeze_panes = ws.cell(row=first_item_row, column=4)
