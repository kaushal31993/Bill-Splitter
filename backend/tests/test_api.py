import io
from decimal import Decimal

from openpyxl import load_workbook


def make_event(client, name="Weekend trip", people=("Me", "Priya", "Raj")):
    client.post("/api/people", json={"name": people[0], "is_owner": True})
    event = client.post("/api/events", json={"name": name}).json()
    ids = {p["display_name"]: p["id"] for p in event["participants"]}
    for person in people[1:]:
        p = client.post(f"/api/events/{event['id']}/participants", json={"name": person}).json()
        ids[p["display_name"]] = p["id"]
    return event["id"], ids


def add_item(client, bill_id, name, total_cents, participant_ids):
    return client.post(
        f"/api/bills/{bill_id}/items",
        json={
            "name": name,
            "quantity": 1,
            "unit_price_cents": total_cents,
            "total_cents": total_cents,
            "participant_ids": participant_ids,
        },
    ).json()


class TestHealthAndConfig:
    def test_health(self, client):
        assert client.get("/api/health").json() == {"status": "ok"}

    def test_config_reports_extraction_availability(self, client):
        body = client.get("/api/config").json()
        assert "extraction_enabled" in body
        assert body["currency"] == "USD"


class TestPeople:
    def test_owner_is_unique(self, client):
        client.post("/api/people", json={"name": "Me", "is_owner": True})
        client.post("/api/people", json={"name": "Someone Else", "is_owner": True})
        owner = client.get("/api/people/owner").json()
        assert owner["name"] == "Someone Else"
        owners = [p for p in client.get("/api/people").json() if p["is_owner"]]
        assert len(owners) == 1

    def test_names_are_reused_not_duplicated(self, client):
        a = client.post("/api/people", json={"name": "Priya"}).json()
        b = client.post("/api/people", json={"name": "priya"}).json()
        assert a["id"] == b["id"]

    def test_owner_404_before_setup(self, client):
        assert client.get("/api/people/owner").status_code == 404


class TestEvents:
    def test_owner_joins_every_event_roster(self, client):
        client.post("/api/people", json={"name": "Me", "is_owner": True})
        event = client.post("/api/events", json={"name": "Dinner"}).json()
        assert [p["display_name"] for p in event["participants"]] == ["Me"]
        assert event["participants"][0]["is_owner"] is True

    def test_roster_is_per_event(self, client):
        e1, ids1 = make_event(client, "Trip", ("Me", "Priya"))
        e2, ids2 = make_event(client, "Dinner", ("Me", "Raj"))
        names1 = {p["display_name"] for p in client.get(f"/api/events/{e1}").json()["participants"]}
        names2 = {p["display_name"] for p in client.get(f"/api/events/{e2}").json()["participants"]}
        assert names1 == {"Me", "Priya"}
        assert names2 == {"Me", "Raj"}

    def test_duplicate_participant_rejected(self, client):
        event_id, _ = make_event(client)
        r = client.post(f"/api/events/{event_id}/participants", json={"name": "Priya"})
        assert r.status_code == 409

    def test_rename_participant_is_event_scoped(self, client):
        event_id, ids = make_event(client)
        r = client.patch(
            f"/api/events/{event_id}/participants/{ids['Raj']}",
            json={"display_name": "Raj K."},
        )
        assert r.json()["display_name"] == "Raj K."

    def test_missing_event_404(self, client):
        assert client.get("/api/events/9999").status_code == 404

    def test_delete_event(self, client):
        event_id, _ = make_event(client)
        assert client.delete(f"/api/events/{event_id}").status_code == 204
        assert client.get(f"/api/events/{event_id}").status_code == 404


class TestParticipantRemovalGuards:
    def test_blocked_while_assigned_to_an_item(self, client):
        event_id, ids = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        add_item(client, bill["id"], "Pasta", 1800, [ids["Priya"]])

        r = client.delete(f"/api/events/{event_id}/participants/{ids['Priya']}")
        assert r.status_code == 409
        assert "Reassign" in r.json()["detail"]

    def test_blocked_while_payer(self, client):
        event_id, ids = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        client.patch(f"/api/bills/{bill['id']}", json={"payer_id": ids["Raj"]})

        r = client.delete(f"/api/events/{event_id}/participants/{ids['Raj']}")
        assert r.status_code == 409
        assert "payer" in r.json()["detail"]

    def test_allowed_when_unencumbered(self, client):
        event_id, ids = make_event(client)
        assert client.delete(f"/api/events/{event_id}/participants/{ids['Raj']}").status_code == 204


class TestBillsAndItems:
    def test_bill_defaults_to_owner_as_payer(self, client):
        event_id, ids = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        assert bill["payer_id"] == ids["Me"]
        assert bill["label"] == "Bill 1"

    def test_bills_are_labelled_in_sequence(self, client):
        event_id, _ = make_event(client)
        labels = [
            client.post(f"/api/events/{event_id}/bills", json={}).json()["label"]
            for _ in range(3)
        ]
        assert labels == ["Bill 1", "Bill 2", "Bill 3"]

    def test_total_derives_from_quantity_and_unit_price(self, client):
        event_id, ids = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        item = client.post(
            f"/api/bills/{bill['id']}/items",
            json={"name": "Beers", "quantity": 3, "unit_price_cents": 650},
        ).json()
        assert item["total_cents"] == 1950

    def test_assignment_to_foreign_participant_rejected(self, client):
        event_a, ids_a = make_event(client, "A", ("Me", "Priya"))
        event_b, ids_b = make_event(client, "B", ("Me", "Raj"))
        bill = client.post(f"/api/events/{event_a}/bills", json={}).json()
        r = client.post(
            f"/api/bills/{bill['id']}/items",
            json={"name": "x", "total_cents": 100, "participant_ids": [ids_b["Raj"]]},
        )
        assert r.status_code == 422

    def test_bulk_assign_everyone(self, client):
        event_id, ids = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        for i in range(3):
            add_item(client, bill["id"], f"item {i}", 900, [])

        everyone = list(ids.values())
        updated = client.post(
            f"/api/bills/{bill['id']}/assign", json={"participant_ids": everyone}
        ).json()
        for item in updated["items"]:
            assert sorted(item["participant_ids"]) == sorted(everyone)

    def test_bulk_assign_only_unassigned_preserves_exceptions(self, client):
        event_id, ids = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        private = add_item(client, bill["id"], "Private", 500, [ids["Raj"]])
        add_item(client, bill["id"], "Shared", 900, [])

        client.post(
            f"/api/bills/{bill['id']}/assign",
            json={"participant_ids": list(ids.values()), "only_unassigned": True},
        )
        items = client.get(f"/api/events/{event_id}").json()["bills"][0]["items"]
        by_name = {i["name"]: i for i in items}
        assert by_name["Private"]["participant_ids"] == [ids["Raj"]]
        assert len(by_name["Shared"]["participant_ids"]) == 3

    def test_replacing_assignments(self, client):
        event_id, ids = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        item = add_item(client, bill["id"], "Wine", 4000, list(ids.values()))
        updated = client.put(
            f"/api/items/{item['id']}/assignments",
            json={"participant_ids": [ids["Me"], ids["Priya"]]},
        ).json()
        assert sorted(updated["participant_ids"]) == sorted([ids["Me"], ids["Priya"]])

    def test_delete_item(self, client):
        event_id, ids = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        item = add_item(client, bill["id"], "Oops", 100, [])
        assert client.delete(f"/api/items/{item['id']}").status_code == 204
        assert client.get(f"/api/events/{event_id}").json()["bills"][0]["items"] == []

    def test_negative_tax_rejected(self, client):
        event_id, _ = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        assert client.patch(f"/api/bills/{bill['id']}", json={"tax_cents": -1}).status_code == 422

    def test_upload_rejects_empty_file(self, client):
        event_id, _ = make_event(client)
        r = client.post(
            f"/api/events/{event_id}/bills/upload",
            files={"file": ("receipt.jpg", b"", "image/jpeg")},
        )
        assert r.status_code == 400

    def test_upload_survives_extraction_failure(self, client):
        """A junk upload must still produce a usable, empty bill — losing the
        upload because extraction failed would be the worst outcome."""
        event_id, _ = make_event(client)
        r = client.post(
            f"/api/events/{event_id}/bills/upload",
            files={"file": ("receipt.jpg", b"not really a jpeg", "image/jpeg")},
        )
        assert r.status_code == 201
        bill = r.json()
        assert bill["extraction_status"] == "failed"
        assert bill["extraction_error"]
        assert bill["items"] == []


class TestTotals:
    def test_end_to_end_totals(self, client):
        event_id, ids = make_event(client)
        me, priya, raj = ids["Me"], ids["Priya"], ids["Raj"]

        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        add_item(client, bill["id"], "Shared platter", 3000, [me, priya, raj])
        add_item(client, bill["id"], "Raj's steak", 3000, [raj])
        client.patch(
            f"/api/bills/{bill['id']}",
            json={"tax_cents": 600, "tip_cents": 1200, "payer_id": me},
        )

        totals = client.get(f"/api/events/{event_id}/totals").json()
        assert totals["grand_total_cents"] == 6000 + 600 + 1200
        assert sum(totals["totals_cents"].values()) == totals["grand_total_cents"]
        assert totals["is_complete"] is True

        # Raj ordered more, so carries more of the tax and tip.
        share = {s["participant_id"]: s for s in totals["bills"][0]["shares"]}
        assert share[raj]["tax_cents"] > share[priya]["tax_cents"]
        assert share[raj]["tip_cents"] > share[priya]["tip_cents"]

        debts = {(d["from_participant_id"], d["to_participant_id"]): d["amount_cents"]
                 for d in totals["debts"]}
        assert set(debts) == {(priya, me), (raj, me)}

    def test_unassigned_item_marks_event_incomplete(self, client):
        event_id, ids = make_event(client)
        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        item = add_item(client, bill["id"], "Mystery", 500, [])
        totals = client.get(f"/api/events/{event_id}/totals").json()
        assert totals["is_complete"] is False
        assert totals["bills"][0]["unassigned_item_ids"] == [item["id"]]

    def test_two_payers_net_per_pair(self, client):
        event_id, ids = make_event(client, people=("Me", "Priya"))
        me, priya = ids["Me"], ids["Priya"]

        b1 = client.post(f"/api/events/{event_id}/bills", json={}).json()
        add_item(client, b1["id"], "Dinner", 6000, [me, priya])
        client.patch(f"/api/bills/{b1['id']}", json={"payer_id": me})

        b2 = client.post(f"/api/events/{event_id}/bills", json={}).json()
        add_item(client, b2["id"], "Cab", 2000, [me, priya])
        client.patch(f"/api/bills/{b2['id']}", json={"payer_id": priya})

        totals = client.get(f"/api/events/{event_id}/totals").json()
        assert len(totals["debts"]) == 1
        debt = totals["debts"][0]
        assert (debt["from_participant_id"], debt["to_participant_id"]) == (priya, me)
        assert debt["amount_cents"] == 2000


class TestExport:
    def test_workbook_structure_and_live_formulas(self, client):
        event_id, ids = make_event(client, "Team dinner")
        me, priya, raj = ids["Me"], ids["Priya"], ids["Raj"]

        bill = client.post(f"/api/events/{event_id}/bills", json={}).json()
        add_item(client, bill["id"], "Pizza", 2400, [me, priya, raj])
        add_item(client, bill["id"], "Beer", 900, [raj])
        client.patch(
            f"/api/bills/{bill['id']}", json={"tax_cents": 300, "tip_cents": 600, "payer_id": me}
        )

        r = client.get(f"/api/events/{event_id}/export.xlsx")
        assert r.status_code == 200
        assert "team-dinner" in r.headers["content-disposition"]

        wb = load_workbook(io.BytesIO(r.content))
        assert wb.sheetnames[0] == "Summary"
        assert "Bill 1" in wb.sheetnames

        summary = wb["Summary"]
        assert summary["A1"].value == "Team dinner"

        sheet = wb["Bill 1"]
        # People are columns, items are rows.
        header_row = next(
            r_
            for r_ in range(1, sheet.max_row + 1)
            if sheet.cell(row=r_, column=1).value == "Item"
        )
        headers = [sheet.cell(row=header_row, column=c).value for c in range(1, 7)]
        assert headers[:3] == ["Item", "Qty", "Line total"]
        assert set(headers[3:6]) == {"Me", "Priya", "Raj"}

        # Each person's column carries their share of each item.
        raj_col = headers.index("Raj") + 1
        beer_row = next(
            r_
            for r_ in range(header_row + 1, sheet.max_row + 1)
            if sheet.cell(row=r_, column=1).value == "Beer"
        )
        # Beer was Raj's alone, and the cell is a real number, not a string.
        assert sheet.cell(row=beer_row, column=raj_col).value == Decimal("9.00")

        # Column totals are live formulas, not baked-in numbers.
        formulas = [
            sheet.cell(row=r_, column=3).value
            for r_ in range(1, sheet.max_row + 1)
            if isinstance(sheet.cell(row=r_, column=3).value, str)
            and sheet.cell(row=r_, column=3).value.startswith("=SUM")
        ]
        assert formulas, "expected live SUM formulas in the line-total column"

    def test_export_of_empty_event(self, client):
        event_id, _ = make_event(client, "Empty")
        r = client.get(f"/api/events/{event_id}/export.xlsx")
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.sheetnames == ["Summary"]
